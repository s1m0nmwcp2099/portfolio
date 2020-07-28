from openpyxl import *
from tkinter import *

#globally declare wb and sheet variable

#open the existing excel file
wb = load_workbook("spreadsheet.xlsx")

#create the sheet object
sheet = wb.active
def excel():
    #resize the width columns in excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    #write given data to spreadsheet at given locn
    sheet.cell(row = 1, column = 1).value = "Name"
    sheet.cell(row = 1, column = 2).value = "Course"
    sheet.cell(row = 1, column = 3).value = "Semester"
    sheet.cell(row = 1, column = 4).value = "Form Number"
    sheet.cell(row = 1, column = 5).value = "Contact Number"
    sheet.cell(row = 1, column = 6).value = "Email id"
    sheet.cell(row = 1, column = 7).value = "Address"

#function to set focus (cursor)
def focus1(event):
    #set focus on course_field box
    course_field.focus_set()

def focus2(event):
    sem_field.focus_set()

def focus3(event):
    form_no_field.focus_set()

def focus4(event):
    contact_no_field.focus_set()

def focus5(event):
    email_id_field.focus_set()

def focus6(event):
    address_field.focus_set()

#fn for clearing boxes
def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

#fn to write data from gui boxes to spreadsheet
def insert():
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):

        print("Empty input")
    else:
        #assigning max row and column value upto which data is written in spreadsheet to variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        #get method returns current text as string to go to spreadsheet at particular locn
        sheet.cell(row = current_row + 1, column = 1).value = name_field.get()
        sheet.cell(row = current_row + 1, column = 2).value = course_field.get()
        sheet.cell(row = current_row + 1, column = 3).value = sem_field.get()
        sheet.cell(row = current_row + 1, column = 4).value = form_no_field.get()
        sheet.cell(row = current_row + 1, column = 5).value = contact_no_field.get()
        sheet.cell(row = current_row + 1, column = 6).value = email_id_field.get()
        sheet.cell(row = current_row + 1, column = 7).value = address_field.get()

        #save the file
        wb.save("spreadsheet.xlsx")

        #set focus on name_field box
        name_field.focus_set()

        #call the clear() fn
        clear()

#driver code
if __name__ == "__main__":
    #create gui window
    root = Tk()
    #set background colour
    root.configure(background = 'light green')
    #set title
    root.title("registration form")
    #set config of gui window
    root.geometry("500x300")

    excel()

    #create form label
    heading = Label(root, text = "Form", bg = "light green")
    #create name label
    name = Label(root, text = "Name", bg = "light green")
    #create course label
    course = Label(root, text = "Course", bg = "light green")
    #create semester label
    sem = Label(root, text = "Semester", bg = "light green")
    #create form no label
    form_no = Label(root, text = "Form No.", bg = "light green")
    #create contact no label
    contact_no = Label(root, text = "Contact no", bg = "light green")
    #create email id label
    email_id = Label(root, text = "Email id", bg = "light green")
    #create address label
    address = Label(root, text = "Address", bg = "light green")

    #grid method to place widgets at respective posns in table like structure
    heading.grid(row = 0, column = 1)
    name.grid(row = 1, column = 0)
    course.grid(row = 2, column = 0)
    sem.grid(row = 3, column = 0)
    form_no.grid(row = 4, column = 0)
    contact_no.grid(row = 5, column = 0)
    email_id.grid(row = 6, column = 0)
    address.grid(row = 7, column = 0)

    #create text entry box
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    #bind (when enter is pressed, then focus1, 2, 3...)
    name_field.bind("<Return>", focus1)
    course_field.bind("<Return>", focus2)
    sem_field.bind("<Return>", focus3)
    form_no_field.bind("<Return>", focus4)
    contact_no_field.bind("<Return>", focus5)
    email_id_field.bind("<Return>", focus6)

    #grid method to place widgets at respective posns in table like structure
    name_field.grid(row = 1, column = 1, ipadx = "100")
    course_field.grid(row = 2, column = 1, ipadx = "100")
    sem_field.grid(row = 3, column = 1, ipadx = "100")
    form_no_field.grid(row = 4, column = 1, ipadx = "100")
    contact_no_field.grid(row = 5, column = 1, ipadx = "100")
    email_id_field.grid(row = 6, column = 1, ipadx = "100")
    address_field.grid(row = 7, column = 1, ipadx = "100")

    #call excel fn
    excel()

    #create submit button
    submit = Button(root, text = "Submit", fg = "Black", bg = "Red", command = insert)
    submit.grid(row = 8, column = 1)

    #start gui
    root.mainloop()